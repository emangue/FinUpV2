"""
Processador para Fatura BTG Pactual - XLSX com senha

DEPLOY → ProjetoFinancasV5
  Destino: app/domains/upload/processors/raw/excel/btg_fatura_xlsx.py  (NOVO arquivo)
  Registry: adicionar entradas para ('btg pactual', 'fatura', 'excel') e aliases

FORMATO DO ARQUIVO:
  - Arquivo criptografado com senha (padrão: CPF sem pontos/traço)
  - Aba: 'Titular'
  - Colunas: (vazia) | Data | Descrição | (vazia) | Valor | Tipo de compra | Código | Final Cartão
  - Data já vem como datetime Python (sem necessidade de parsing)
  - Valores já em BRL (internacionais já convertidos)

RETORNO: Tuple[List[RawTransaction], BalanceValidation]
  - service.py trata nativamente via isinstance(result, tuple)
  - saldo_inicial=0.0 (fatura, sem saldo inicial)
  - saldo_final = valor da linha "Total da Fatura" no XLSX
  - is_valid = True se soma bate com o total (tolerância 1 centavo)

DEPENDÊNCIAS (adicionar ao requirements.txt se não existirem):
  - msoffcrypto-tool
  - openpyxl
"""

import io
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

import msoffcrypto
import openpyxl

from ..base import RawTransaction, BalanceValidation, PasswordRequiredException

logger = logging.getLogger(__name__)


def process_btg_fatura_xlsx(
    file_path: Path,
    nome_arquivo: str,
    nome_cartao: str = "BTG Pactual",
    final_cartao: str = None,
    senha: Optional[str] = None,
) -> Tuple[List[RawTransaction], BalanceValidation]:
    """
    Processa fatura BTG Pactual em formato XLSX (com ou sem senha).

    Assinatura compatível com o registry do V5:
        processor(file_path, nome_arquivo, nome_cartao, final_cartao)
    O parâmetro `senha` deve ser injetado via functools.partial no registry:
        partial(process_btg_fatura_xlsx, senha=os.getenv('BTG_SENHA'))

    Args:
        file_path: Caminho do arquivo .xlsx
        nome_arquivo: Nome original do arquivo (usado para extrair mes_fatura)
        nome_cartao: Nome do cartão para identificação (padrão: 'BTG Pactual')
        final_cartao: Final do cartão — ignorado aqui, extraído do próprio XLSX por linha
        senha: Senha de proteção do arquivo (CPF sem pontos/traço)

    Returns:
        Tupla (List[RawTransaction], BalanceValidation).
        service.py detecta a tupla via isinstance e salva o balance_validation.

    Raises:
        FileNotFoundError: Se o arquivo não existir.
        ValueError: Se não encontrar a aba ou estrutura esperada.
        Exception: Se a senha estiver errada ou o arquivo for inválido.
    """
    logger.info(f"Processando fatura BTG XLSX: {nome_arquivo}" + (" [com senha]" if senha else ""))

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    # Descriptografar se tiver senha
    wb = _open_workbook(file_path, senha)

    # Localizar aba de transações
    sheet_name = _find_sheet(wb)
    ws = wb[sheet_name]
    logger.debug(f"Aba encontrada: '{sheet_name}'")

    # Localizar linha de cabeçalho e extrair transações
    rows = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows)
    if header_idx is None:
        raise ValueError(f"Cabeçalho 'Data | Descrição | Valor' não encontrado em '{sheet_name}'")

    logger.debug(f"Cabeçalho na linha {header_idx + 1}")

    # Capturar total declarado no documento (linha 'Total da Fatura')
    balance = BalanceValidation()
    balance.saldo_inicial = 0.0  # Fatura não tem saldo inicial
    balance.saldo_final = _extract_total_fatura(rows)
    if balance.saldo_final is not None:
        logger.debug(f"Total declarado na fatura: R$ {balance.saldo_final:.2f}")
    else:
        logger.warning("Total da fatura não encontrado no XLSX")

    mes_fatura = _extract_mes_fatura(nome_arquivo)
    data_criacao = datetime.now()
    transactions: List[RawTransaction] = []

    for row in rows[header_idx + 1:]:
        # Estrutura: (None, Data, Descrição, vazia, Valor, Tipo, Código, Final Cartão)
        if len(row) < 8:
            continue

        _, data_val, descricao, _, valor_val, tipo_compra, codigo, final_cartao_xlsx = row[:8]

        # Ignorar linhas sem data ou valor
        if data_val is None or valor_val is None:
            continue

        # Converter data
        if isinstance(data_val, datetime):
            data_str = data_val.strftime("%Y-%m-%d")
        elif isinstance(data_val, str):
            data_str = _parse_date_str(data_val)
        else:
            logger.warning(f"Formato de data desconhecido: {data_val!r} — linha ignorada")
            continue

        # Limpar descrição
        descricao = str(descricao).strip() if descricao else ""
        if not descricao:
            continue

        # Valor já em BRL (internacionais já convertidos pelo BTG)
        try:
            valor = float(valor_val)
        except (TypeError, ValueError):
            logger.warning(f"Valor inválido: {valor_val!r} — linha ignorada")
            continue

        # Final do cartão como string (vem do próprio XLSX por linha)
        final_cartao_str = str(int(final_cartao_xlsx)) if final_cartao_xlsx is not None else ""

        transactions.append(
            RawTransaction(
                banco="BTG",
                tipo_documento="fatura",
                nome_arquivo=nome_arquivo,
                data_criacao=data_criacao,
                data=data_str,
                lancamento=descricao,
                valor=valor,
                nome_cartao=nome_cartao,
                final_cartao=final_cartao_str,
                mes_fatura=mes_fatura,
            )
        )

    # Validar soma vs total declarado
    balance.soma_transacoes = round(sum(t.valor for t in transactions), 2)
    balance.validate()

    logger.info(f"✅ Fatura BTG XLSX processada: {len(transactions)} transações")
    logger.info(
        f"{'✅' if balance.is_valid else '⚠️ ATENÇÃO'} Validação: "
        f"total declarado R$ {balance.saldo_final:.2f} | "
        f"soma extraída R$ {balance.soma_transacoes:.2f} | "
        f"diferença R$ {balance.diferenca:.2f}"
    )
    return transactions, balance


# ─── Helpers privados ──────────────────────────────────────────────────────────

def _extract_total_fatura(rows: list) -> Optional[float]:
    """
    Localiza e retorna o valor 'Total da Fatura' nas linhas do XLSX.
    """
    for row in rows:
        row_str = [str(c).strip().lower() if c is not None else "" for c in row]
        if any("total da fatura" in c or "total a pagar" in c for c in row_str):
            for cell in row:
                if isinstance(cell, (int, float)) and cell > 0:
                    return float(cell)
    return None


def _open_workbook(file_path: Path, senha: Optional[str]) -> openpyxl.Workbook:
    """Abre workbook descriptografando com msoffcrypto se necessário."""
    # Verificar se o arquivo é criptografado independentemente de ter senha
    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        is_encrypted = office_file.is_encrypted()

    if is_encrypted and not senha:
        raise PasswordRequiredException(filename=file_path.name, wrong_password=False)

    if is_encrypted and senha:
        try:
            with open(file_path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=senha)
                decrypted = io.BytesIO()
                office_file.decrypt(decrypted)
            return openpyxl.load_workbook(decrypted, read_only=True, data_only=True)
        except Exception as e:
            # Senha incorreta
            raise PasswordRequiredException(filename=file_path.name, wrong_password=True) from e

    return openpyxl.load_workbook(file_path, read_only=True, data_only=True)


def _find_sheet(wb: openpyxl.Workbook) -> str:
    """Retorna o nome da aba principal de transações."""
    candidates = ["Titular", "Sheet1", "Planilha1"]
    for name in candidates:
        if name in wb.sheetnames:
            return name
    return wb.sheetnames[0]


def _find_header_row(rows: list) -> Optional[int]:
    """Encontra o índice da linha de cabeçalho com 'Data' e 'Descrição'."""
    for i, row in enumerate(rows):
        row_str = [str(c).strip().lower() if c is not None else "" for c in row]
        if "data" in row_str and ("descrição" in row_str or "descricao" in row_str):
            return i
    return None


def _parse_date_str(date_str: str) -> str:
    """Converte string de data para YYYY-MM-DD, tentando vários formatos."""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    logger.warning(f"Não foi possível parsear data: {date_str!r}")
    return date_str


def _extract_mes_fatura(nome_arquivo: str) -> str:
    """
    Extrai mês da fatura do nome do arquivo.
    Ex: '2026-02-01_Fatura_..._BTG.xlsx' → '202602'
    """
    match = re.search(r'(\d{4})-(\d{2})-\d{2}', nome_arquivo)
    if match:
        return match.group(1) + match.group(2)
    match = re.search(r'(\d{6})', nome_arquivo)
    if match:
        return match.group(1)
    return datetime.now().strftime("%Y%m")
